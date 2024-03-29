import openpyxl
from datetime import datetime


def get_row_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path) # to load excel file
    sheet = workbook[sheet_name] # to retrieve sheet - enter sheet name
    return sheet.max_row


def get_column_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path) # to load excel file
    sheet = workbook[sheet_name] # to retrieve sheet - enter sheet name
    return sheet.max_column


def read_data(path,sheet_name,row_number,column_number):
    workbook = openpyxl.load_workbook(path)   # to load excel file
    sheet = workbook[sheet_name]  # to retrieve sheet - enter sheet name
    return sheet.cell(row=row_number,column=column_number).value


def write_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)  # to load excel file
    sheet = workbook[sheet_name]  # to retrieve sheet - enter sheet name
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


